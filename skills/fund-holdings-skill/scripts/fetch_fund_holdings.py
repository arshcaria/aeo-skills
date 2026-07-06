"""
Fetch complete holding history for any Chinese mutual fund
from eastmoney (天天基金) API and output to Excel.

Only 半年报 (6/30) and 年报 (12/31) — the periods with complete holdings disclosure.
Includes fund total AUM from gmbd endpoint for asset composition analysis.

Usage:
    python fetch_fund_holdings.py                          # interactive
    python fetch_fund_holdings.py --code 004685           # specify fund code
    python fetch_fund_holdings.py --code 004685 --output ./out.xlsx
    python fetch_fund_holdings.py --code 004685 --start-year 2020

Requirements:
    pip install openpyxl
"""

import urllib.request
import urllib.error
import re
import json
import time
import sys
import os
import argparse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Referer': 'https://fundf10.eastmoney.com/'
}


# ═══════════════════════════════════════════════════════
#  Network layer
# ═══════════════════════════════════════════════════════

def fetch_raw(url, timeout=30):
    """Fetch raw bytes from URL."""
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def fetch_holdings_html(fund_code, year):
    """Fetch fund holdings HTML for a given year."""
    url = (f'https://fundf10.eastmoney.com/FundArchivesDatas.aspx'
           f'?type=jjcc&code={fund_code}&topline=500&year={year}&month=&rt={time.time()}')
    return fetch_raw(url).decode('utf-8')


def fetch_fund_aum(fund_code):
    """Fetch fund AUM (期末净资产) history from gmbd endpoint."""
    url = (f'https://fundf10.eastmoney.com/FundArchivesDatas.aspx'
           f'?type=gmbd&code={fund_code}')
    raw = fetch_raw(url).decode('utf-8')

    m = re.search(r'gmbd_apidata=\{ content:"(.+?)"\s*}', raw, re.DOTALL)
    if not m:
        return {}
    content = m.group(1).replace('\\"', '"')

    aum_map = {}
    rows = re.findall(r'<tr>(.*?)</tr>', content, re.DOTALL)
    for row_html in rows:
        cells = re.findall(r'<td[^>]*>(.*?)</td>', row_html, re.DOTALL)
        cleaned = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
        if len(cleaned) >= 5 and cleaned[0]:
            date = cleaned[0]
            net_assets_str = cleaned[4]  # 期末净资产（亿元）
            try:
                net_assets = float(net_assets_str.replace(',', ''))
                aum_map[date] = net_assets
            except ValueError:
                pass
    return aum_map


def fetch_fund_name(fund_code):
    """Fetch fund name from the holdings page."""
    try:
        html = fetch_holdings_html(fund_code, 2025)
        inner = extract_content(html)
        m = re.search(r"<a[^>]*title='([^']+)'", inner)
        if m:
            return m.group(1)
    except Exception:
        pass
    return fund_code


# ═══════════════════════════════════════════════════════
#  Parsing layer
# ═══════════════════════════════════════════════════════

def extract_content(html_text):
    """Extract the inner HTML content from the JS variable `var apidata={ content:"..." }`."""
    content_match = re.search(r'content\s*:\s*"', html_text)
    if not content_match:
        return html_text
    start = content_match.end()

    content_parts = []
    i = start
    while i < len(html_text):
        ch = html_text[i]
        if ch == '\\':
            if i + 1 < len(html_text):
                next_ch = html_text[i + 1]
                if next_ch in ['"', '\\', '/', 'b', 'f', 'n', 'r', 't']:
                    content_parts.append(next_ch if next_ch != 'n' else '\n')
                    i += 2
                    continue
                elif next_ch == 'u':
                    hex_str = html_text[i + 2:i + 6]
                    try:
                        content_parts.append(chr(int(hex_str, 16)))
                    except ValueError:
                        content_parts.append('?')
                    i += 6
                    continue
            content_parts.append(ch)
            i += 1
        elif ch == '"':
            break
        else:
            content_parts.append(ch)
            i += 1
    return ''.join(content_parts)


def parse_periods(html):
    """Parse HTML content into list of (period_label, rows)."""
    periods = []
    h4_blocks = re.split(r'<h4\s+class=[\'"]t[\'"]>', html)

    for block in h4_blocks[1:]:
        h4_match = re.search(r'(.*?)</h4>', block, re.DOTALL)
        if not h4_match:
            continue
        label = h4_match.group(1)
        label = re.sub(r'<[^>]+>', '', label).strip()
        label = re.sub(r'\s+', ' ', label)

        table_bodies = re.findall(r'<tbody>(.*?)</tbody>', block, re.DOTALL)
        all_rows = []
        for tbody in table_bodies:
            rows = re.findall(r'<tr>(.*?)</tr>', tbody, re.DOTALL)
            for row_html in rows:
                cells = re.findall(r'<td[^>]*>(.*?)</td>', row_html, re.DOTALL)
                cleaned = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
                if len(cleaned) >= 7:
                    all_rows.append(cleaned)
        if all_rows:
            periods.append((label, all_rows))
    return periods


def clean_label(label):
    """Parse period label into (date_str, type_str)."""
    label_clean = re.sub(r'\s+', '', label)
    m = re.search(r'(\d{4})年(\d)季度', label_clean)
    if m:
        y, q = m.group(1), m.group(2)
        qt_map = {'1': ('03-31', '一季报'), '2': ('06-30', '中报/半年报'),
                   '3': ('09-30', '三季报'), '4': ('12-31', '年报')}
        d, t = qt_map.get(q, ('', ''))
        return f"{y}-{d}", t
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', label_clean)
    if m:
        d = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        if d.endswith('12-31'):
            t = '年报'
        elif d.endswith('06-30'):
            t = '中报/半年报'
        elif d.endswith('09-30'):
            t = '三季报'
        elif d.endswith('03-31'):
            t = '一季报'
        else:
            t = '其他'
        return d, t
    return label_clean, "未知"


def normalize_rows(periods):
    """Convert parsed periods into structured (date_str, type_str, [stock_rows])."""
    result = []
    for label, rows in periods:
        date_str, ptype = clean_label(label)
        data_rows = []
        for row in rows:
            if len(row) < 5:
                continue
            seq = row[0].strip()
            code = row[1].strip()
            name = row[2].strip() if len(row) > 2 else ""
            ratio_str = row[4].strip() if len(row) > 4 else ""
            shares_str = row[5].strip() if len(row) > 5 else ""
            mv_str = row[6].strip() if len(row) > 6 else ""

            if not seq.isdigit() or not code:
                continue

            try:
                ratio = float(ratio_str.replace('%', '').replace(',', ''))
            except ValueError:
                ratio = 0.0
            try:
                shares = float(shares_str.replace(',', ''))
            except ValueError:
                shares = 0.0
            try:
                mv = float(mv_str.replace(',', ''))
            except ValueError:
                mv = 0.0

            data_rows.append({
                'seq': int(seq),
                'code': code,
                'name': name,
                'ratio': round(ratio, 4),
                'shares': round(shares, 2),
                'market_value': round(mv, 2)
            })

        if data_rows:
            data_rows.sort(key=lambda x: x['seq'])
            result.append((date_str, ptype, data_rows))

    result.sort(key=lambda x: x[0], reverse=True)
    return result


# ═══════════════════════════════════════════════════════
#  Excel generation
# ═══════════════════════════════════════════════════════

def create_excel(all_data, aum_map, fund_code, fund_name, output_file):
    """Create Excel workbook — only 年报 & 半年报, one sheet per period."""
    wb = openpyxl.Workbook()

    # ── Styles ──
    hdr_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(horizontal="center", vertical="center")

    annual_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    semi_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header_row(ws, row, headers, col_widths):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = thin
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    def write_row(ws, row, values, fill=None):
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = data_font
            c.alignment = data_align
            c.border = thin
            if fill:
                c.fill = fill

    # ═══ 持仓概览 ═══
    ws = wb.active
    ws.title = "持仓概览"

    ws.merge_cells('A1:L1')
    c = ws['A1']
    c.value = f"{fund_name} ({fund_code}) 历年持仓概览"
    c.font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:L2')
    c = ws['A2']
    c.value = ("数据来源：天天基金 | 仅含半年报(6/30)和年报(12/31)完整持仓 | "
               "持仓市值占比 = 个股持仓市值 / 股票总持仓市值 | "
               "其他资产 = 基金总规模 - 股票持仓")
    c.font = Font(name="微软雅黑", size=9, color="888888")
    c.alignment = Alignment(horizontal="center")

    ov_headers = [
        "报告期", "报告类型", "股票数",
        "基金总规模\n(亿元)",
        "股票持仓\n(亿元)",
        "股票占比\n(%)",
        "其他资产\n(亿元)",
        "其他占比\n(%)",
        "第一大重仓", "第一大占比(%)*\n(占股票持仓)",
        "前十大占比(%)*\n(占股票持仓)"
    ]
    ov_widths = [16, 12, 8, 14, 14, 12, 14, 12, 18, 16, 16]
    style_header_row(ws, 4, ov_headers, ov_widths)

    ov_row = 5
    for date_str, ptype, rows in all_data:
        n = len(rows)
        total_mv_wan = sum(r['market_value'] for r in rows)
        total_mv_yi = round(total_mv_wan / 10000, 4)

        aum_yi = aum_map.get(date_str, 0)

        other_yi = round(aum_yi - total_mv_yi, 4) if aum_yi > 0 else 0
        stock_pct = round(total_mv_yi / aum_yi * 100, 1) if aum_yi > 0 else 0
        other_pct = round(other_yi / aum_yi * 100, 1) if aum_yi > 0 else 0

        sorted_rows = sorted(rows, key=lambda x: x['market_value'], reverse=True)
        top1 = sorted_rows[0]
        top1_pct = round(top1['market_value'] / total_mv_wan * 100, 2) if total_mv_wan > 0 else 0
        top10_pct = round(
            sum(r['market_value'] for r in sorted_rows[:10]) / total_mv_wan * 100, 2
        ) if total_mv_wan > 0 else 0

        vals = [date_str, ptype, n,
                aum_yi, total_mv_yi, stock_pct, other_yi, other_pct,
                top1['name'], top1_pct, top10_pct]
        fill = annual_fill if '年报' in ptype else semi_fill
        write_row(ws, ov_row, vals, fill)
        ov_row += 1

    # ═══ One sheet per period ═══
    period_headers = ["序号", "股票代码", "股票名称",
                      "占净值比例(%)",
                      "持仓市值占比(%)",
                      "持股数(万股)", "持仓市值(万元)"]
    period_widths = [8, 14, 18, 15, 15, 15, 18]

    for date_str, ptype, rows in all_data:
        safe_type = ptype.replace('/', '-').replace('\\', '-')
        sheet_name = f"{date_str} {safe_type}"[:31]
        ws_p = wb.create_sheet(sheet_name)

        total_mv_wan = sum(r['market_value'] for r in rows)
        total_mv_yi = round(total_mv_wan / 10000, 4)
        aum_yi = aum_map.get(date_str, 0)
        stock_pct = round(total_mv_yi / aum_yi * 100, 1) if aum_yi > 0 else 0

        sorted_rows = sorted(rows, key=lambda x: x['market_value'], reverse=True)

        ws_p.merge_cells('A1:G1')
        ws_p.cell(row=1, column=1,
                  value=(f"{fund_name}  {date_str}  [{ptype}]  {len(rows)}只股票  "
                         f"基金规模{aum_yi}亿 | 股票持仓{total_mv_yi}亿({stock_pct}%)")).font = \
            Font(name="微软雅黑", size=12, bold=True, color="2F5496")
        ws_p.row_dimensions[1].height = 28

        style_header_row(ws_p, 3, period_headers, period_widths)

        dr = 4
        fill = annual_fill if '年报' in ptype else semi_fill
        for new_seq, r in enumerate(sorted_rows, 1):
            orig_ratio = round(r['ratio'], 3)
            calc_ratio = round(r['market_value'] / total_mv_wan * 100, 3) if total_mv_wan > 0 else 0
            write_row(ws_p, dr,
                      [new_seq, r['code'], r['name'], orig_ratio, calc_ratio,
                       r['shares'], r['market_value']],
                      fill)
            dr += 1

    wb.save(output_file)


# ═══════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='获取公募基金历年半年报/年报完整持仓')
    parser.add_argument('--code', type=str, default='', help='基金代码，如 004685')
    parser.add_argument('--output', type=str, default='', help='输出文件路径')
    parser.add_argument('--start-year', type=int, default=2018, help='起始年份（默认2018）')
    parser.add_argument('--end-year', type=int, default=None, help='结束年份（默认当前年）')
    args = parser.parse_args()

    fund_code = args.code.strip()
    if not fund_code:
        fund_code = input('请输入基金代码（如 004685）: ').strip()

    if args.end_year:
        end_year = args.end_year
    else:
        end_year = time.localtime().tm_year

    if args.output:
        output_file = args.output
    else:
        output_file = f'{fund_code}_历史持仓.xlsx'

    print("=" * 60)
    print(f"  基金持仓数据获取工具")
    print(f"  基金代码: {fund_code}")
    print("=" * 60)

    # ── Fetch fund name ──
    fund_name = fetch_fund_name(fund_code)
    print(f"  基金名称: {fund_name}")

    # ── Fetch fund AUM ──
    print("  获取基金规模数据 (gmbd) ...")
    aum_map = fetch_fund_aum(fund_code)
    print(f"  获取到 {len(aum_map)} 个季度的规模数据")

    # ── Fetch holdings ──
    all_data = []
    years = list(range(args.start_year, end_year + 1))

    for year in years:
        sys.stdout.write(f"  获取 {year} 年持仓数据 ... ")
        sys.stdout.flush()
        try:
            html = fetch_holdings_html(fund_code, year)
            inner = extract_content(html)
            periods_raw = parse_periods(inner)
            periods = normalize_rows(periods_raw)
            # Filter: only 年报 and 中报/半年报
            periods = [(d, t, r) for d, t, r in periods if t in ('年报', '中报/半年报')]
            n_periods = len(periods)
            n_stocks = sum(len(r) for _, _, r in periods)
            print(f"{n_periods}个, {n_stocks}条")
            all_data.extend(periods)
        except Exception as e:
            print(f"失败 ({e})")
        time.sleep(0.3)

    all_data.sort(key=lambda x: x[0], reverse=True)

    print(f"\n  总计: {len(all_data)} 个报告期（仅年报+半年报）")
    for d, t, r in all_data:
        aum = aum_map.get(d, 0)
        total_mv = sum(x['market_value'] for x in r)
        print(f"    {d} [{t}] {len(r)}只  基金{aum}亿  股票{round(total_mv / 10000, 2)}亿")

    print(f"\n  生成 Excel 文件 ...")
    create_excel(all_data, aum_map, fund_code, fund_name, output_file)
    print(f"  已保存: {output_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
